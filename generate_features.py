"""
Genera el Excel de funcionalidades del proyecto workflow-doc-studio_back.
Ejecutar: python generate_features.py
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Datos ──────────────────────────────────────────────────────────────────────

FEATURES = [
    # ── PIPELINE GENERAL ──────────────────────────────────────────────────────
    (1,  "Renderizado de PDF desde templateJson",           "Implementado",  "Recibe un JSON con páginas y elementos y devuelve bytes PDF válidos. Punto de entrada: POST /render."),
    (2,  "Renderizado de PDF desde workflow (DAG)",         "Implementado",  "Recibe un grafo de nodos/aristas, lo ejecuta y luego renderiza. Punto de entrada: POST /generate."),
    (3,  "Smoke test sin archivos",                         "Implementado",  "GET /test-render genera un PDF mínimo en memoria para verificar que el servicio funciona."),
    (4,  "CLI de línea de comandos",                        "Implementado",  "run.py acepta subcomandos: render (archivo a PDF), server (levanta FastAPI), test-render."),
    (5,  "API REST con FastAPI",                            "Implementado",  "Endpoints: POST /generate, POST /render, GET /health, GET /test-render. Swagger UI automático en /docs."),
    (6,  "Servicio sin estado (stateless)",                 "Implementado",  "Sin base de datos, sin caché, sin persistencia. Cada request es independiente y autocontenido."),

    # ── WORKFLOW / EJECUCIÓN DE GRAFOS ────────────────────────────────────────
    (7,  "Ordenamiento topológico de nodos (Kahn)",         "Implementado",  "workflow/executor.py ordena y ejecuta los nodos del DAG garantizando el orden correcto de dependencias."),
    (8,  "Nodo webhook-trigger",                            "Implementado",  "Inyecta valores por defecto al contexto de datos según el schema declarado en el nodo."),
    (9,  "Nodo data-processor",                             "Implementado",  "Aplica mappings de campos con transforms: uppercase, lowercase, capitalize, trim."),
    (10, "Nodo document-designer",                          "Implementado",  "Nodo terminal que contiene el templateJson a renderizar."),
    (11, "Nodo data-viewer",                                "Implementado",  "Nodo de inspección de solo lectura, no modifica el contexto de datos."),
    (12, "Compatibilidad workflow sin wrapper",             "Implementado",  "POST /generate acepta el workflow en la raíz o envuelto en key 'workflow' para compatibilidad con callers antiguos."),

    # ── SISTEMA DE COORDENADAS ────────────────────────────────────────────────
    (13, "Conversión mm → puntos (pt)",                     "Implementado",  "1 mm = 2.8346 pt. Toda medida del JSON pasa por coordinate.py antes de llegar a ReportLab."),
    (14, "Inversión del eje Y (top-left → bottom-left)",    "Implementado",  "El JSON usa Y creciente hacia abajo (CSS). ReportLab usa Y creciente hacia arriba. coordinate.py hace la conversión."),
    (15, "Función safe_float para valores CSS",             "Implementado",  "Convierte strings CSS ('normal', '12px', 'auto') a float de forma segura, devolviendo el default si no es convertible."),

    # ── PÁGINAS ───────────────────────────────────────────────────────────────
    (16, "Páginas múltiples",                               "Implementado",  "El template puede tener N páginas; cada una se renderiza como showPage() en el canvas de ReportLab."),
    (17, "Páginas de diferentes tamaños",                   "Implementado",  "Cada página define su propio width/height en mm; se aplican individualmente al canvas."),
    (18, "Fondo de página (color sólido)",                  "Implementado",  "Si la página tiene fill de tipo solid se pinta el fondo antes de renderizar los elementos."),
    (19, "Visibilidad condicional de páginas",              "Implementado",  "Páginas con visible:false son omitidas del PDF final."),
    (20, "Márgenes de página",                              "Implementado",  "apply_margin() en coordinate.py calcula el rectángulo útil descontando márgenes en mm."),

    # ── NORMALIZACIÓN DE TEMPLATE ─────────────────────────────────────────────
    (21, "Normalización de template (strip UI)",            "Implementado",  "normalize.py elimina campos solo-UI: createdAt, zIndex, locked, pagePreview, elementCount, etc."),
    (22, "Índice de assets (O(1) lookup)",                  "Implementado",  "normalize.py construye asset_index: {assetId → asset} para acceso directo sin recorrer el árbol."),
    (23, "Índice de content areas (O(1) lookup)",           "Implementado",  "normalize.py aplana el árbol recursivo de áreas y construye area_index: {areaId → area}."),
    (24, "Colapso de borders/fills desactivados",           "Implementado",  "Borders con enabled:false y fills de tipo 'none' se normalizan a None para simplificar los renderers."),

    # ── ELEMENTOS: FORMAS ─────────────────────────────────────────────────────
    (25, "Rectángulo",                                      "Implementado",  "shape_renderer.py dibuja rectángulo sólido con fill y border opcionales."),
    (26, "Rectángulo con bordes redondeados",               "Implementado",  "Si border.radius.unified > 0 usa canvas.roundRect(); soporta radio en mm."),
    (27, "Elipse / círculo",                                "Implementado",  "shape_renderer.py dibuja elipse; si width == height produce círculo perfecto."),
    (28, "Triángulo (apuntando hacia arriba)",              "Implementado",  "shape_renderer.py dibuja triángulo isósceles con apex arriba usando beginPath/moveTo/lineTo."),
    (29, "Fill sólido en formas",                           "Implementado",  "Aplica color y opacidad (canal alpha en ReportLab Color) al relleno de la forma."),
    (30, "Fill tipo gradiente en formas",                   "Pendiente",     "El código tiene el placeholder 'implement when needed' pero no está implementado. ReportLab soporta linearGradient."),

    # ── ELEMENTOS: IMAGEN ─────────────────────────────────────────────────────
    (31, "Imagen desde asset del template",                 "Implementado",  "image_renderer.py busca el asset en asset_index por assetId y resuelve la ruta con ASSETS_BASE_PATH."),
    (32, "Imagen desde URL/path directo",                   "Implementado",  "kind='url' usa directamente la ruta del campo source.url."),
    (33, "Imagen placeholder",                              "Implementado",  "Cuando la imagen no existe o es kind='placeholder' dibuja un rectángulo gris con texto '[ Image ]'."),
    (34, "Fit mode: contain",                               "Implementado",  "preserveAspectRatio=True mantiene la relación de aspecto dentro del bounding box."),
    (35, "Fit mode: cover / fill",                          "Implementado",  "preserveAspectRatio=False estira la imagen para llenar el bounding box exacto."),
    (36, "Rotación de imagen",                              "Implementado",  "Rota alrededor del centro del elemento usando translate/rotate/translate en el canvas."),
    (37, "Imagen desde base64",                             "Pendiente",     "No hay soporte para data URIs (data:image/png;base64,...). Requeriría decodificar y escribir a archivo temporal."),
    (38, "Imagen SVG",                                      "Pendiente",     "ReportLab no renderiza SVG nativamente. Requeriría svglib o cairosvg como dependencia adicional."),

    # ── ELEMENTOS: TEXT ───────────────────────────────────────────────────────
    (39, "Caja de texto fija (type: text)",                 "Implementado",  "text_renderer.py renderiza texto plano (sin HTML) en posición fija con estilos inline del elemento."),
    (40, "Alineación vertical en text (top/middle/bottom)", "Implementado",  "Calcula el offset vertical según la altura del párrafo renderizado vs. la altura del elemento."),
    (41, "Padding interno en text",                         "Implementado",  "paddingTop/Right/Bottom/Left en mm definen el espacio interior del elemento text."),
    (42, "Transformación de texto (upper/lower/capitalize)","Implementado",  "text_transform del TextStyle se aplica al contenido antes de pasar a ReportLab."),

    # ── ELEMENTOS: CONTENT AREA ───────────────────────────────────────────────
    (43, "Content area con HTML rico",                      "Implementado",  "contentarea_renderer.py procesa HTML del editor React y lo convierte a flowables de ReportLab."),
    (44, "Parsing HTML: spans con estilos CSS inline",      "Implementado",  "html_parser.py parsea color, font-weight, font-style, font-size, text-decoration, vertical-align."),
    (45, "Parsing HTML: tags semánticos (b, i, u, em)",     "Implementado",  "Shorcut tags se mapean a InlineStyle con bold/italic/underline."),
    (46, "Parsing HTML: listas (ul, ol, li)",               "Implementado",  "Genera objetos Paragraph con list_item=True, list_depth y list_type (bullet/numbered)."),
    (47, "Parsing HTML: saltos de línea (br)",              "Implementado",  "Emite un TextRun con text='\\n' que se convierte a <br/> en el XML de ReportLab."),
    (48, "Parsing HTML: párrafos y divs",                   "Implementado",  "div/p crean un nuevo Paragraph en la lista de salida."),
    (49, "Variables en content area (var-tag)",             "Implementado",  "Spans con class='var-tag' data-var='nombre' se resuelven contra el contexto de datos."),
    (50, "Variable $pageNumber",                            "Implementado",  "Inyectada por page_renderer al renderizar cada página; se reemplaza en los TextRun de cada área."),
    (51, "Variable $pageCount / $totalPages",               "Implementado",  "Inyectada por page_renderer junto con $pageNumber."),
    (52, "Variable $today (fecha ISO)",                     "Implementado",  "variable_resolver.py la resuelve con date.today().isoformat()."),
    (53, "Variable $now (datetime ISO)",                    "Implementado",  "variable_resolver.py la resuelve con datetime.now().isoformat(timespec='seconds')."),
    (54, "Acceso a campos anidados con dot-notation",       "Implementado",  "variable_resolver._get_nested() soporta 'user.address.city' sobre dicts anidados."),
    (55, "Acceso a índice de array con dot-notation",       "Implementado",  "Segmentos numéricos en el path (ej: 'items.0.name') acceden al elemento de la lista por índice."),
    (56, "Sub-áreas inline (area-tag)",                     "Implementado",  "Spans con class='area-tag' expanden el contenido de otra área en línea sin crear salto de párrafo."),
    (57, "Sub-áreas condicionales (inline-condition)",      "Implementado",  "area_resolver.py evalúa selectionScript o selectionVariable y elige trueAreaId o falseAreaId."),
    (58, "Tablas embebidas en content area (element-tag)",  "Implementado",  "Spans con class='element-tag' data-type='table' insertan una Table platypus como flowable bloque."),
    (59, "KeepInFrame modo shrink en content area",         "Implementado",  "Si el contenido no cabe en el área, ReportLab lo encoge proporcionalmente hasta que entre."),
    (60, "Strips de caracteres invisibles del editor",      "Implementado",  "html_parser._STRIP_INVISIBLE elimina zero-width spaces y soft hyphens insertados por contenteditable."),

    # ── ELEMENTOS: TABLA ─────────────────────────────────────────────────────
    (61, "Tabla con datos dinámicos (dataSource)",          "Implementado",  "table_renderer.py lee una lista de dicts del contexto de datos y genera una fila por ítem."),
    (62, "Tabla con filas explícitas",                      "Implementado",  "body.rows / header.rows / footer.rows permiten definir el contenido directamente en el template."),
    (63, "Header de tabla (repite en cada página)",         "Implementado",  "repeatRows=1 en ReportLab Table hace que el header se repita al paginar."),
    (64, "Footer de tabla",                                 "Implementado",  "footer_cfg con enabled:true agrega filas de pie con el mismo estilo del header."),
    (65, "Anchos de columna en mm",                         "Implementado",  "widthUnit='mm' convierte el valor con mm(). Escala proporcionalmente si la suma difiere del ancho total."),
    (66, "Anchos de columna en porcentaje",                 "Implementado",  "widthUnit='%' calcula total_w_pt * val / 100."),
    (67, "Anchos de columna por proporción (widthRatio)",   "Implementado",  "widthRatio 0-1 multiplica por el ancho total disponible."),
    (68, "Zebra striping (filas alternas)",                 "Implementado",  "alternateRowFill con color sólido aplica BACKGROUND a filas pares del body."),
    (69, "Borde exterior de tabla (BOX)",                   "Implementado",  "tableBorder.unified con enabled:true agrega BOX al TableStyle."),
    (70, "Bordes internos de tabla (INNERGRID)",            "Implementado",  "cellBorder.unified con enabled:true agrega INNERGRID al TableStyle."),
    (71, "Bordes por celda en tabla embebida",              "Implementado",  "En tablas dentro de contentarea, cada celda puede tener border.inline con configuración por lado."),
    (72, "Colspan / rowspan en tablas",                     "Pendiente",     "No hay soporte para fusionar celdas. Requeriría SPAN en TableStyle y ajuste del array de datos."),
    (73, "Tabla con paginación / salto de página",          "Parcial",       "KeepInFrame con mode='shrink' puede encoger la tabla. Pero no hay soporte explícito de SPLIT para cortar la tabla entre páginas."),

    # ── ELEMENTOS: QR Y BARCODE ───────────────────────────────────────────────
    (74, "Código QR",                                       "Implementado",  "qr_renderer.py usa la librería qrcode. Soporta valor estático o dinámico del contexto, colores y niveles de corrección."),
    (75, "Barcode CODE128",                                 "Implementado",  "barcode_renderer.py usa python-barcode. Valor estático o dinámico del contexto de datos."),
    (76, "Barcode EAN-13 / EAN-8",                          "Implementado",  "Simbología configurable: CODE128, CODE39, EAN13, EAN8, ITF."),
    (77, "Mostrar/ocultar texto del barcode",               "Implementado",  "Campo showText controla si se renderiza el texto bajo las barras."),

    # ── ESTILOS ───────────────────────────────────────────────────────────────
    (78, "StyleRegistry con lookup por ID",                 "Implementado",  "style_registry.py construye tablas de lookup para TextStyle, ParagraphStyle, FillStyle, BorderStyle."),
    (79, "TextStyle: familia, peso, tamaño, color",         "Implementado",  "Resuelto a fontName de ReportLab con fallback a Helvetica. Color como hex string."),
    (80, "TextStyle: bold, italic, underline, tachado",     "Implementado",  "Mapeados a tags XML de ReportLab (<b>, <i>, <u>) o a variante de fuente bold/italic."),
    (81, "TextStyle: superscript y subscript",              "Implementado",  "Mapeados a <super> y <sub> en el XML del Paragraph de ReportLab."),
    (82, "TextStyle: letter spacing (repurposed indent)",   "Implementado",  "letter_spacing se usa como leftIndent en el ParagraphStyle (limitación de ReportLab)."),
    (83, "TextStyle: line-height",                          "Implementado",  "Se multiplica por font_size para obtener el leading de ReportLab."),
    (84, "ParagraphStyle: alineación (left/center/right/justify)", "Implementado", "Mapeado a TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY de ReportLab."),
    (85, "ParagraphStyle: indents y espaciado",             "Implementado",  "firstLineIndent, leftIndent, spaceBefore, spaceAfter en mm → pt."),
    (86, "FillStyle: sólido con opacidad",                  "Implementado",  "rl_color() crea Color con canal alpha cuando opacity < 1.0."),
    (87, "BorderStyle: unificado (todas las caras iguales)","Implementado",  "border.mode='unified' dibuja un rect/roundRect con un solo color y ancho."),
    (88, "BorderStyle: por lado (top/right/bottom/left)",   "Implementado",  "border.mode='sides' dibuja cada lado con su propio color y ancho."),
    (89, "BorderStyle: radio de esquinas",                  "Implementado",  "radius_x en mm se convierte a pt y se usa en roundRect."),
    (90, "BorderStyle: márgenes internos",                  "Implementado",  "margin_top/right/bottom/left reducen el rectángulo de borde hacia adentro."),

    # ── FUENTES ───────────────────────────────────────────────────────────────
    (91, "Fuentes bundled (directorio fonts/)",             "Implementado",  "FontManager.load_directory() escanea fonts/ del proyecto en cada request."),
    (92, "Fuentes del sistema operativo",                   "Implementado",  "load_system_fonts() detecta Windows/macOS/Linux y escanea los directorios de fuentes del OS."),
    (93, "Fuentes declaradas en el template",               "Implementado",  "templateJson.fonts[] con family + variants (path, weight, italic) se registran con mayor prioridad."),
    (94, "Fallback a Helvetica si fuente no existe",        "Implementado",  "font_manager.resolve() cae a _BUILTIN (Helvetica/Times/Courier) si la familia no está registrada."),
    (95, "Detección bold/italic por nombre de archivo",     "Implementado",  "_parse_font_filename() detecta variantes por keywords: bold, italic, semibold, oblique, etc."),
    (96, "Fuentes Google Fonts",                            "Pendiente",     "No hay descarga automática de Google Fonts. El usuario debe incluir los .ttf en fonts/ manualmente."),

    # ── API / INFRAESTRUCTURA ─────────────────────────────────────────────────
    (97, "Variable de entorno ASSETS_BASE_PATH",            "Implementado",  "Prefijo del filesystem para rutas de assets de imagen. Permite despliegue con volúmenes montados."),
    (98, "Health check",                                    "Implementado",  "GET /health devuelve {status: 'ok'} para liveness probes de Kubernetes/Docker."),
    (99, "Swagger UI auto-generado",                        "Implementado",  "FastAPI genera /docs con Swagger UI y /redoc automáticamente."),
    (100,"Metadata PDF (título, autor, asunto)",            "Pendiente",     "ReportLab soporta canvas.setTitle/setAuthor/setSubject pero no hay campos en el schema del template."),
    (101,"Contraseña / encriptación del PDF",               "Pendiente",     "ReportLab soporta canvas.encrypt() para contraseñas de usuario y propietario, pero no está expuesto."),
    (102,"PDF/A compliance",                                "Pendiente",     "No hay soporte para PDF/A (archivo a largo plazo). Requeriría gestión de color ICC y metadatos XMP."),
    (103,"Hipervínculos en texto",                          "Pendiente",     "ReportLab soporta canvas.linkURL() y <a href=...> en Paragraph XML, pero no hay campo en el schema."),
    (104,"Marcadores / bookmarks en PDF",                   "Pendiente",     "canvas.bookmarkPage() y canvas.addOutlineEntry() de ReportLab no están expuestos en el schema."),
    (105,"Watermark / marca de agua",                       "Pendiente",     "Podría implementarse como una capa transparente de texto o imagen en cada página, pero no existe."),
    (106,"Sombras en elementos",                            "Pendiente",     "ReportLab no tiene sombras nativas. Requeriría dibujar un rectángulo desplazado y semi-transparente."),
    (107,"Soporte RTL (árabe, hebreo)",                     "Pendiente",     "wordWrap='CJK' cubre CJK pero no RTL. Requeriría bidi algorithm (python-bidi) y fuente RTL."),
    (108,"Streaming de respuesta (chunked)",                "Pendiente",     "La respuesta devuelve los bytes completos. Para PDFs grandes podría beneficiarse de StreamingResponse."),
    (109,"Validación del schema del template (Pydantic)",   "Pendiente",     "models/template.py existe pero no se usa para validar la entrada en los endpoints /generate y /render."),
    (110,"Autenticación / autorización en la API",          "Pendiente",     "No hay API key, JWT ni ningún mecanismo de autenticación. El servicio es completamente abierto."),
    (111,"Rate limiting",                                   "Pendiente",     "No hay límite de requests por IP ni por cliente. Necesario si se expone a internet."),
    (112,"Logging estructurado (JSON)",                     "Pendiente",     "No hay logging configurado. Errores llegan solo como excepciones no capturadas al stderr."),
    (113,"Caché de PDFs generados",                         "Pendiente",     "Por diseño el servicio es stateless, pero un caché Redis con hash del input ahorraría re-renders."),
    (114,"Exportación a DOCX",                              "Pendiente",     "No hay soporte para Word. Requeriría python-docx como motor alternativo."),
    (115,"Exportación a HTML",                              "Pendiente",     "No hay renderizado a HTML. Podría ser útil para preview rápido sin ReportLab."),
    (116,"Exportación a PNG / JPG (preview de página)",     "Pendiente",     "No hay rasterización. Requeriría pdf2image o pypdfium2 como post-procesado."),
    (117,"Gráficos / charts en PDF",                        "Pendiente",     "No hay tipo de elemento chart. ReportLab tiene reportlab.graphics.charts pero no está integrado."),
    (118,"Listas numeradas con numeración automática",      "Parcial",       "html_parser detecta ol/li y marca list_type='numbered', pero _make_list_item usa bulletText='' (sin número visible). Requiere ListFlowable de ReportLab."),
    (119,"Fill gradiente en content area / text",           "Pendiente",     "Solo solid fill está implementado en contentarea y text_renderer. El tipo 'gradient' es ignorado."),
    (120,"Overflow / clipping de elementos",                "Pendiente",     "Si el contenido es mayor que el elemento, KeepInFrame lo encoge pero no lo corta. No hay modo clip explícito."),
]

# ── Colores ────────────────────────────────────────────────────────────────────
COLOR_HEADER_BG   = "1E3A5F"   # azul oscuro
COLOR_HEADER_FONT = "FFFFFF"
COLOR_IMPL_BG     = "D6F5D6"   # verde claro
COLOR_IMPL_FONT   = "1A5C1A"
COLOR_PEND_BG     = "FFF3CC"   # amarillo claro
COLOR_PEND_FONT   = "7A5200"
COLOR_PART_BG     = "FFE0CC"   # naranja claro
COLOR_PART_FONT   = "7A2D00"
COLOR_ROW_ALT     = "F5F8FF"   # azul muy tenue para filas alternas

THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def status_style(status: str):
    if status == "Implementado":
        return PatternFill("solid", fgColor=COLOR_IMPL_BG), Font(color=COLOR_IMPL_FONT, bold=True)
    if status == "Pendiente":
        return PatternFill("solid", fgColor=COLOR_PEND_BG), Font(color=COLOR_PEND_FONT, bold=True)
    return PatternFill("solid", fgColor=COLOR_PART_BG), Font(color=COLOR_PART_FONT, bold=True)


def build_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Funcionalidades"

    # ── Título principal ──────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "Workflow Doc Studio — Inventario de Funcionalidades"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color=COLOR_HEADER_FONT)
    title_cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Fila de encabezados ───────────────────────────────────────────────────
    headers = ["ID", "Funcionalidad", "Estado", "Descripción"]
    col_widths = [6, 48, 16, 90]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color=COLOR_HEADER_FONT)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 22

    # ── Filas de datos ────────────────────────────────────────────────────────
    for data_row in FEATURES:
        row_num = data_row[0] + 2   # offset: fila 1 = título, fila 2 = headers
        id_, func, estado, desc = data_row

        id_cell   = ws.cell(row=row_num, column=1, value=id_)
        func_cell = ws.cell(row=row_num, column=2, value=func)
        stat_cell = ws.cell(row=row_num, column=3, value=estado)
        desc_cell = ws.cell(row=row_num, column=4, value=desc)

        # Alineación
        id_cell.alignment   = Alignment(horizontal="center", vertical="top")
        func_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        stat_cell.alignment = Alignment(horizontal="center", vertical="top")
        desc_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Fuente base
        for cell in (id_cell, func_cell, desc_cell):
            cell.font = Font(name="Calibri", size=9)

        # Color del estado
        bg_fill, status_font = status_style(estado)
        stat_cell.fill = bg_fill
        stat_cell.font = status_font

        # Fila alterna (solo columnas sin color de estado)
        if id_ % 2 == 0:
            alt_fill = PatternFill("solid", fgColor=COLOR_ROW_ALT)
            for cell in (id_cell, func_cell, desc_cell):
                cell.fill = alt_fill

        # Bordes
        for cell in (id_cell, func_cell, stat_cell, desc_cell):
            cell.border = BORDER

        ws.row_dimensions[row_num].height = 30

    # ── Panel de resumen (debajo de la tabla) ─────────────────────────────────
    total = len(FEATURES)
    impl  = sum(1 for f in FEATURES if f[2] == "Implementado")
    pend  = sum(1 for f in FEATURES if f[2] == "Pendiente")
    part  = sum(1 for f in FEATURES if f[2] == "Parcial")

    summary_row = total + 4
    ws.merge_cells(f"A{summary_row}:D{summary_row}")
    s = ws.cell(row=summary_row, column=1,
                value=f"Resumen:  ✅ Implementado: {impl}   ⏳ Parcial: {part}   🔲 Pendiente: {pend}   |   Total: {total} funcionalidades")
    s.font = Font(name="Calibri", size=10, bold=True, color=COLOR_HEADER_FONT)
    s.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[summary_row].height = 22

    # ── Freeze panes ──────────────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    # ── Auto-filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A2:D{total + 2}"

    # Guardar
    out = "/home/user/workflow-doc-studio_back/funcionalidades_workflow_doc_studio.xlsx"
    wb.save(out)
    print(f"Excel generado: {out}")
    print(f"  ✅ Implementado: {impl}")
    print(f"  ⏳ Parcial:      {part}")
    print(f"  🔲 Pendiente:    {pend}")
    print(f"  Total:           {total}")


if __name__ == "__main__":
    build_excel()
